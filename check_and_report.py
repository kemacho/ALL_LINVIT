import os
import glob
import threading
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.formatting.rule import CellIsRule
from PySide6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, QLabel,
                               QLineEdit, QPushButton, QComboBox, QProgressBar, QMessageBox, QFileDialog,
                               QTextEdit, QDialog, QFrame)
from PySide6.QtCore import Qt, Signal, QObject

# Названия подпапок СИ
nameSI = ['0 Заявка и приложение', '1 Распоряжение по заявке',
          '2 Решение по заявке', '3 Заключения по ОМТД',
          '4 Акт выбора ПК', '4.1 Направление-заявка',
          '5 Протоколы СИ', '6 Заключение СИ',
          '7 Программа АСП', '8 Акт АСП',
          '9 Распоряжение на анализ', '10 Решение о выдаче',
          '11 Сертификат', '12 Доп.материалы']

# Названия подпапок ИК
nameIK = ['1 Распоряжение', '2 Извещение',
          '3 Программа ИК', '4 Акт по ОМТД',
          '5 Акт выбора ПК', '5.1 Направление-заявка',
          '6 Протоколы ИК', '7 Акт по испытаниям ИК',
          '8 Программа АСП', '9 Акт АСП',
          '10 Акт по результатам ИК', '11 Решение по ИК',
          '12 Доп. материалы']

SI_TEMPLATES = {
    "Общий шаблон": ['1', '1', '1', '1', '1', '1', 'Any', '1', '1', '1', '1', '1', '2', 'Any'],
    "Шаблон РЖД": ['2', '1', '1', '1', '1', '1', 'Any', '1', '1', '1', '1', '2', '2', 'Any']
}

IK_TEMPLATES = {
    "Общий шаблон": ['1', '1', '1', '1', '1', '1', 'Any', '1', '1', '1', '1', '1', 'Any'],
    "Шаблон РЖД": ['1', '1', '1', '1', '1', '1', 'Any', '1', '1', '1', '1', '1', 'Any']
}


class WarningDialog(QDialog):
    """Диалог для отображения предупреждений"""

    def __init__(self, warnings, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Предупреждения")
        self.setMinimumSize(500, 300)

        layout = QVBoxLayout(self)

        self.text_edit = QTextEdit()
        self.text_edit.setReadOnly(True)
        self.text_edit.setPlainText(warnings)

        layout.addWidget(QLabel("Обратите внимание на следующие папки:"))
        layout.addWidget(self.text_edit)

        close_button = QPushButton("Закрыть")
        close_button.clicked.connect(self.accept)
        layout.addWidget(close_button)


class WorkerSignals(QObject):
    progress = Signal(int)
    message = Signal(str)
    finished = Signal(str)
    error = Signal(str)
    warnings = Signal(list)  # Новый сигнал для предупреждений


class FolderProcessor(threading.Thread):
    def __init__(self, inpath, fileqnt, fileqnt_ik):
        super().__init__()
        self.inpath = inpath
        self.fileqnt = fileqnt
        self.fileqnt_ik = fileqnt_ik
        self.signals = WorkerSignals()
        self.warnings = []  # Сохраняем предупреждения

    def run(self):
        try:
            self.process_folders()
            excel_path = self.create_excel_report()
            self.signals.finished.emit(excel_path)
        except Exception as e:
            self.signals.error.emit(str(e))

    def add_warning(self, pt, nm1, nm2):
        try:
            # Берем путь относительно корневой папки
            relative_path = os.path.relpath(pt, self.inpath)
            self.warnings.append(
                f'Проверь: {relative_path}, там находится {nm1} файла, вместо {nm2}')
        except:
            # Если не получается вычислить относительный путь, используем только имя папки
            folder_name = os.path.basename(pt)
            self.warnings.append(f'Пожалуйста проверьте папку: {folder_name}, там находится {nm1} файла, вместо {nm2}')

    def process_folder_with_one_file(self, old_folder, contents, Pos_dest, Neg_dest):
        if len(contents) == 1 and old_folder != Pos_dest:
            os.rename(old_folder, Pos_dest)
        elif len(contents) == 0 and old_folder != Neg_dest:
            os.rename(old_folder, Neg_dest)
        elif len(contents) > 1:
            os.rename(old_folder, Pos_dest)
            self.add_warning(old_folder, len(contents), '1')

    def process_folder_with_four_files(self, old_folder, contents, Pos_dest, Neg_dest):
        if len(contents) == 4 and old_folder != Pos_dest:
            os.rename(old_folder, Pos_dest)
        elif len(contents) < 4 and old_folder != Neg_dest:
            os.rename(old_folder, Neg_dest)
        elif len(contents) > 4:
            os.rename(old_folder, Pos_dest)
            self.add_warning(old_folder, len(contents), '4')

    def process_folder_with_any_files(self, old_folder, contents, Pos_dest, Neg_dest):
        if len(contents) > 0 and old_folder != Pos_dest:
            os.rename(old_folder, Pos_dest)
        elif len(contents) == 0 and old_folder != Neg_dest:
            os.rename(old_folder, Neg_dest)
        elif len(contents) > 20:
            os.rename(old_folder, Pos_dest)
            self.add_warning(old_folder, len(contents), 'возможно не так много')

    def process_folder_with_two_files(self, old_folder, contents, Pos_dest, Norm_dest, Neg_dest):
        if len(contents) == 2 and old_folder != Pos_dest:
            os.rename(old_folder, Pos_dest)
        elif len(contents) == 1 and old_folder != Norm_dest:
            os.rename(old_folder, Norm_dest)
        elif len(contents) == 0 and old_folder != Neg_dest:
            os.rename(old_folder, Neg_dest)
        elif len(contents) > 2:
            os.rename(old_folder, Pos_dest)
            self.add_warning(old_folder, len(contents), '2')

    def check(self, FileQNT, old_folder, contents, Pos_dest, Norm_dest, Neg_dest):
        filtered_contents = [f for f in contents if f != "Thumbs.db"]

        if FileQNT == "1":
            self.process_folder_with_one_file(old_folder, filtered_contents, Pos_dest, Neg_dest)
        if FileQNT == "2":
            self.process_folder_with_two_files(old_folder, filtered_contents, Pos_dest, Norm_dest, Neg_dest)
        if FileQNT == "4":
            self.process_folder_with_four_files(old_folder, filtered_contents, Pos_dest, Neg_dest)
        elif FileQNT == 'Any':
            self.process_folder_with_any_files(old_folder, filtered_contents, Pos_dest, Neg_dest)

    def process_folders(self):
        try:
            folder_names = os.listdir(self.inpath)
        except Exception as e:
            self.signals.error.emit(f"Ошибка при чтении папки: {e}")
            return

        total_folders = len(folder_names)
        for i, name in enumerate(folder_names):
            pathSI = os.path.join(self.inpath, name, '0. СИ')
            pathIK1 = os.path.join(self.inpath, name, '1. ИК-1')
            pathIK2 = os.path.join(self.inpath, name, '2. ИК-2')

            # Проверка для папки СИ
            for j in range(len(self.fileqnt)):
                nameSIzv = nameSI[j] + '*'
                try:
                    old_folder = glob.glob(os.path.join(pathSI, nameSIzv))
                    if old_folder:
                        old_folder = old_folder[0]
                        contents = os.listdir(old_folder)

                        folder = os.path.join(pathSI, nameSI[j])
                        Pos_dest = str(folder) + ' (+)'
                        Norm_dest = str(folder) + ' (+—)'
                        Neg_dest = str(folder) + ' (—)'

                        self.check(self.fileqnt[j], old_folder, contents, Pos_dest, Norm_dest, Neg_dest)
                except Exception as e:
                    print(f"Ошибка при обработке {nameSI[j]} в {name}: {e}")

            # Проверка для папки ИК1
            for j in range(len(self.fileqnt_ik)):
                nameIK1zv = nameIK[j] + '*'
                try:
                    old_folder = glob.glob(os.path.join(pathIK1, nameIK1zv))
                    if old_folder:
                        old_folder = old_folder[0]
                        contents = os.listdir(old_folder)

                        folder = os.path.join(pathIK1, nameIK[j])
                        Pos_dest = str(folder) + ' (+)'
                        Norm_dest = str(folder) + ' (+—)'
                        Neg_dest = str(folder) + ' (—)'

                        self.check(self.fileqnt_ik[j], old_folder, contents, Pos_dest, Norm_dest, Neg_dest)
                except Exception as e:
                    print(f"Ошибка при обработке {nameIK[j]} в {name} (ИК1): {e}")

            # Проверка для папки ИК2
            for j in range(len(self.fileqnt_ik)):
                nameIK2zv = nameIK[j] + '*'
                try:
                    old_folder = glob.glob(os.path.join(pathIK2, nameIK2zv))
                    if old_folder:
                        old_folder = old_folder[0]
                        contents = os.listdir(old_folder)

                        folder = os.path.join(pathIK2, nameIK[j])
                        Pos_dest = str(folder) + ' (+)'
                        Norm_dest = str(folder) + ' (+—)'
                        Neg_dest = str(folder) + ' (—)'

                        self.check(self.fileqnt_ik[j], old_folder, contents, Pos_dest, Norm_dest, Neg_dest)
                except Exception as e:
                    print(f"Ошибка при обработке {nameIK[j]} в {name} (ИК2): {e}")

            progress = int((i + 1) / total_folders * 100)
            self.signals.progress.emit(progress)
            self.signals.message.emit(f"Обработано папок: {i + 1} из {total_folders}")

        # Отправляем предупреждения через сигнал
        if self.warnings:
            self.signals.warnings.emit(self.warnings)

    def create_excel_report(self):
        from openpyxl.styles import PatternFill
        from openpyxl.formatting.rule import FormulaRule

        wb = Workbook()
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        sheets = {
            'СИ': nameSI,
            'ИК-1': nameIK,
            'ИК-2': nameIK
        }

        # Определяем цвета
        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        yellow_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        gray_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

        for sheet_name, headers in sheets.items():
            ws = wb.create_sheet(title=sheet_name)
            ws.column_dimensions['A'].width = 50
            for col in range(2, len(headers) + 2):
                ws.column_dimensions[chr(64 + col)].width = 15

            headers_with_name = ['Название папки'] + headers
            ws.row_dimensions[1].height = 30

            for col_num, header in enumerate(headers_with_name, start=1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')

        # Сначала заполняем данные
        folder_names = os.listdir(self.inpath)
        for folder_name in folder_names:
            pathSI = os.path.join(self.inpath, folder_name, '0. СИ')
            if os.path.exists(pathSI):
                self.process_folder_for_excel(wb['СИ'], folder_name, pathSI, nameSI)

            pathIK1 = os.path.join(self.inpath, folder_name, '1. ИК-1')
            if os.path.exists(pathIK1):
                self.process_folder_for_excel(wb['ИК-1'], folder_name, pathIK1, nameIK)

            pathIK2 = os.path.join(self.inpath, folder_name, '2. ИК-2')
            if os.path.exists(pathIK2):
                self.process_folder_for_excel(wb['ИК-2'], folder_name, pathIK2, nameIK)

        # Теперь добавляем условное форматирование
        for sheet_name in sheets.keys():
            ws = wb[sheet_name]

            # Определяем диапазон с данными
            max_row = ws.max_row
            if max_row <= 1:  # Только заголовок
                continue

            # Начинаем с колонки B (2), заканчиваем последней колонкой с данными
            start_col = 2
            end_col = len(nameSI) + 1 if sheet_name == 'СИ' else len(nameIK) + 1

            # Преобразуем в буквы Excel
            start_col_letter = chr(64 + start_col)
            end_col_letter = chr(64 + end_col)

            # Диапазон для форматирования (исключая заголовок)
            data_range = f'{start_col_letter}2:{end_col_letter}{max_row}'


            # Для ячеек со знаком "+"
            formula1 = f'{start_col_letter}2="+"'  # Это шаблон, Excel сам адаптирует для каждой ячейки
            ws.conditional_formatting.add(data_range,
                                          FormulaRule(formula=[f'{start_col_letter}2="+"'],
                                                      fill=green_fill))

            # Для ячеек со знаком "+-"
            ws.conditional_formatting.add(data_range,
                                          FormulaRule(formula=[f'{start_col_letter}2="+-"'],
                                                      fill=yellow_fill))

            # Для ячеек со знаком "-"
            ws.conditional_formatting.add(data_range,
                                          FormulaRule(formula=[f'{start_col_letter}2="-"'],
                                                      fill=red_fill))

            # Для ячеек с текстом "Нет папки"
            ws.conditional_formatting.add(data_range,
                                          FormulaRule(formula=[f'{start_col_letter}2="Нет папки"'],
                                                      fill=gray_fill))

        output_path = os.path.join(self.inpath, "results.xlsx")
        wb.save(output_path)
        return output_path

    def process_folder_for_excel(self, worksheet, folder_name, base_path, subfolder_names):
        row_num = worksheet.max_row + 1 if worksheet.max_row > 1 else 2
        worksheet.cell(row=row_num, column=1, value=folder_name)

        for i, subfolder in enumerate(subfolder_names, start=2):
            subfolder_path = os.path.join(base_path, subfolder + '*')
            matched_folders = glob.glob(subfolder_path)

            if not matched_folders:
                worksheet.cell(row=row_num, column=i, value="Нет папки")
                continue

            folder = matched_folders[0]
            status = self.get_folder_status(folder)
            worksheet.cell(row=row_num, column=i, value=status)

    def get_folder_status(self, folder_path):
        folder_name = os.path.basename(folder_path)
        if ' (+)' in folder_name:
            return '+'
        elif ' (—)' in folder_name:
            return '-'
        elif ' (+—)' in folder_name:
            return '+-'
        else:
            return '?'


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Обработчик папок")
        self.setMinimumSize(1000, 750)  # Увеличил для размещения всех элементов

        self.central_widget = QWidget()
        self.setCentralWidget(self.central_widget)
        self.layout = QVBoxLayout(self.central_widget)

        # Для хранения предупреждений
        self.warnings_list = []

        # Верхняя часть: выбор папки
        self.folder_frame = QFrame()
        self.folder_frame.setFrameStyle(QFrame.Panel | QFrame.Raised)
        self.folder_layout = QHBoxLayout(self.folder_frame)
        self.folder_layout.setContentsMargins(10, 10, 10, 10)

        self.folder_label = QLabel("Корневая папка:")
        self.folder_label.setFixedWidth(100)
        self.folder_entry = QLineEdit()
        self.folder_entry.setPlaceholderText("Выберите папку...")
        self.browse_button = QPushButton("Обзор")
        self.browse_button.setFixedWidth(80)
        self.browse_button.clicked.connect(self.browse_folder)

        self.folder_layout.addWidget(self.folder_label)
        self.folder_layout.addWidget(self.folder_entry)
        self.folder_layout.addWidget(self.browse_button)

        # Центральная часть: две вертикальные области (СИ и ИК)
        self.center_frame = QWidget()
        self.center_layout = QHBoxLayout(self.center_frame)
        self.center_layout.setContentsMargins(5, 5, 5, 5)
        self.center_layout.setSpacing(10)

        # Левая область: СИ
        self.si_frame = QFrame()
        self.si_frame.setFrameStyle(QFrame.Panel | QFrame.Raised)
        self.si_frame.setLineWidth(2)
        self.si_layout = QVBoxLayout(self.si_frame)
        self.si_layout.setContentsMargins(10, 10, 10, 10)
        self.si_layout.setSpacing(5)

        # Заголовок СИ
        si_title = QLabel("СИ (Сертификационные испытания)")
        si_title.setAlignment(Qt.AlignCenter)
        si_title.setStyleSheet("font-weight: bold; font-size: 12pt; margin-bottom: 5px;")
        self.si_layout.addWidget(si_title)

        # Выбор шаблона СИ
        self.template_si_frame = QWidget()
        self.template_si_layout = QHBoxLayout(self.template_si_frame)
        self.template_si_layout.setContentsMargins(0, 0, 0, 10)

        self.template_si_label = QLabel("Шаблон:")
        self.template_si_label.setFixedWidth(60)
        self.template_si_combo = QComboBox()
        self.template_si_combo.addItems(["Пользовательский"] + list(SI_TEMPLATES.keys()))
        self.template_si_combo.setCurrentText("Общий шаблон")
        self.template_si_combo.currentTextChanged.connect(self.apply_si_template)

        self.template_si_layout.addWidget(self.template_si_label)
        self.template_si_layout.addWidget(self.template_si_combo)

        self.si_layout.addWidget(self.template_si_frame)

        # Параметры СИ - статический список
        self.si_params_widget = QWidget()
        self.si_params_layout = QVBoxLayout(self.si_params_widget)
        self.si_params_layout.setContentsMargins(5, 5, 5, 5)
        self.si_params_layout.setSpacing(2)  # Уменьшил расстояние между элементами

        self.si_combos = []
        options = ["1", "2", "4", "Any"]

        for i, name in enumerate(nameSI):
            frame = QWidget()
            frame.setFixedHeight(30)  # Фиксированная высота для выравнивания
            layout = QHBoxLayout(frame)
            layout.setContentsMargins(5, 0, 5, 0)
            layout.setSpacing(10)

            label = QLabel(f"{name}")
            label.setFixedWidth(180)  # Фиксированная ширина для выравнивания
            combo = QComboBox()
            combo.addItems(options)
            combo.setFixedWidth(80)  # Фиксированная ширина комбобоксов

            layout.addWidget(label)
            layout.addStretch()
            layout.addWidget(combo)
            self.si_params_layout.addWidget(frame)
            self.si_combos.append(combo)

        # Добавляем растягивающий элемент в конец для выравнивания
        self.si_params_layout.addStretch()

        self.si_layout.addWidget(self.si_params_widget)

        # Правая область: ИК
        self.ik_frame = QFrame()
        self.ik_frame.setFrameStyle(QFrame.Panel | QFrame.Raised)
        self.ik_frame.setLineWidth(2)
        self.ik_layout = QVBoxLayout(self.ik_frame)
        self.ik_layout.setContentsMargins(10, 10, 10, 10)
        self.ik_layout.setSpacing(5)

        # Заголовок ИК
        ik_title = QLabel("ИК (Инспекционный контроль)")
        ik_title.setAlignment(Qt.AlignCenter)
        ik_title.setStyleSheet("font-weight: bold; font-size: 12pt; margin-bottom: 5px;")
        self.ik_layout.addWidget(ik_title)

        # Выбор шаблона ИК
        self.template_ik_frame = QWidget()
        self.template_ik_layout = QHBoxLayout(self.template_ik_frame)
        self.template_ik_layout.setContentsMargins(0, 0, 0, 10)

        self.template_ik_label = QLabel("Шаблон:")
        self.template_ik_label.setFixedWidth(60)
        self.template_ik_combo = QComboBox()
        self.template_ik_combo.addItems(["Пользовательский"] + list(IK_TEMPLATES.keys()))
        self.template_ik_combo.setCurrentText("Общий шаблон")
        self.template_ik_combo.currentTextChanged.connect(self.apply_ik_template)

        self.template_ik_layout.addWidget(self.template_ik_label)
        self.template_ik_layout.addWidget(self.template_ik_combo)

        self.ik_layout.addWidget(self.template_ik_frame)

        # Параметры ИК - статический список
        self.ik_params_widget = QWidget()
        self.ik_params_layout = QVBoxLayout(self.ik_params_widget)
        self.ik_params_layout.setContentsMargins(5, 5, 5, 5)
        self.ik_params_layout.setSpacing(2)  # Уменьшил расстояние между элементами

        self.ik_combos = []

        for i, name in enumerate(nameIK):
            frame = QWidget()
            frame.setFixedHeight(30)  # Фиксированная высота для выравнивания
            layout = QHBoxLayout(frame)
            layout.setContentsMargins(5, 0, 5, 0)
            layout.setSpacing(10)

            label = QLabel(f"{name}")
            label.setFixedWidth(180)  # Фиксированная ширина для выравнивания
            combo = QComboBox()
            combo.addItems(options)
            combo.setFixedWidth(80)  # Фиксированная ширина комбобоксов

            layout.addWidget(label)
            layout.addStretch()
            layout.addWidget(combo)
            self.ik_params_layout.addWidget(frame)
            self.ik_combos.append(combo)

        # Добавляем растягивающий элемент в конец для выравнивания
        self.ik_params_layout.addStretch()

        self.ik_layout.addWidget(self.ik_params_widget)

        # Добавляем обе области в центральный layout
        self.center_layout.addWidget(self.si_frame)
        self.center_layout.addWidget(self.ik_frame)

        # Нижняя часть: прогресс, сообщения и кнопки
        self.bottom_frame = QFrame()
        self.bottom_layout = QVBoxLayout(self.bottom_frame)
        self.bottom_layout.setContentsMargins(10, 10, 10, 10)

        # Прогресс бар
        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)

        # Сообщение
        self.message_label = QLabel()
        self.message_label.setAlignment(Qt.AlignCenter)
        self.message_label.setFixedHeight(20)

        # Кнопки
        self.buttons_frame = QWidget()
        self.buttons_layout = QHBoxLayout(self.buttons_frame)
        self.buttons_layout.setContentsMargins(0, 10, 0, 0)

        self.start_button = QPushButton("Начать обработку")
        self.start_button.setFixedHeight(35)
        self.start_button.clicked.connect(self.start_processing)

        self.warnings_button = QPushButton("Показать предупреждения")
        self.warnings_button.setFixedHeight(35)
        self.warnings_button.clicked.connect(self.show_warnings)
        self.warnings_button.setEnabled(False)

        self.buttons_layout.addWidget(self.start_button)
        self.buttons_layout.addWidget(self.warnings_button)

        self.bottom_layout.addWidget(self.progress_bar)
        self.bottom_layout.addWidget(self.message_label)
        self.bottom_layout.addWidget(self.buttons_frame)

        # Добавляем все виджеты в основной layout
        self.layout.addWidget(self.folder_frame)
        self.layout.addWidget(self.center_frame, 1)  # Растягиваем центральную часть
        self.layout.addWidget(self.bottom_frame)

        # Применяем шаблоны по умолчанию
        self.apply_si_template()
        self.apply_ik_template()

    def browse_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "Выберите корневую папку")
        if folder:
            self.folder_entry.setText(folder)

    def apply_si_template(self):
        template_name = self.template_si_combo.currentText()
        if template_name != "Пользовательский":
            template = SI_TEMPLATES.get(template_name)
            if template:
                for i, value in enumerate(template):
                    if i < len(self.si_combos):
                        self.si_combos[i].setCurrentText(value)

    def apply_ik_template(self):
        template_name = self.template_ik_combo.currentText()
        if template_name != "Пользовательский":
            template = IK_TEMPLATES.get(template_name)
            if template:
                for i, value in enumerate(template):
                    if i < len(self.ik_combos):
                        self.ik_combos[i].setCurrentText(value)

    def start_processing(self):
        folder_path = self.folder_entry.text()
        if not folder_path:
            QMessageBox.critical(self, "Ошибка", "Необходимо выбрать корневую папку.")
            return

        fileqnt = [combo.currentText() for combo in self.si_combos]
        fileqnt_ik = [combo.currentText() for combo in self.ik_combos]

        self.progress_bar.setValue(0)
        self.message_label.setText("Подготовка к обработке...")
        self.start_button.setEnabled(False)
        self.warnings_button.setEnabled(False)
        self.warnings_list = []

        self.worker = FolderProcessor(folder_path, fileqnt, fileqnt_ik)
        self.worker.signals.progress.connect(self.update_progress)
        self.worker.signals.message.connect(self.update_message)
        self.worker.signals.finished.connect(self.processing_finished)
        self.worker.signals.error.connect(self.show_error)
        self.worker.signals.warnings.connect(self.save_warnings)
        self.worker.start()

    def update_progress(self, value):
        self.progress_bar.setValue(value)

    def update_message(self, message):
        self.message_label.setText(message)

    def save_warnings(self, warnings):
        self.warnings_list = warnings
        if warnings:
            self.warnings_button.setEnabled(True)

    def processing_finished(self, excel_path):
        self.start_button.setEnabled(True)
        message = f"Обработка завершена!\nОтчет сохранен в:\n{excel_path}"
        if self.warnings_list:
            message += "\n\nЕсть предупреждения. Нажмите 'Показать предупреждения' для просмотра."
        QMessageBox.information(self, "Готово", message)

    def show_warnings(self):
        if self.warnings_list:
            warnings_text = "\n".join(self.warnings_list)
            dialog = WarningDialog(warnings_text, self)
            dialog.exec()

    def show_error(self, error_msg):
        self.start_button.setEnabled(True)
        QMessageBox.critical(self, "Ошибка", error_msg)


if __name__ == "__main__":
    app = QApplication([])
    window = MainWindow()
    window.show()
    app.exec()